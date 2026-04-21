from rest_framework.parsers          import MultiPartParser, FormParser, JSONParser
from rest_framework                  import generics, permissions, status
from rest_framework.decorators       import api_view, permission_classes
from rest_framework.response         import Response
from rest_framework.authtoken.models import Token
from rest_framework.pagination       import PageNumberPagination
from django.db.models                import Sum, Count
from django.db.models.functions      import TruncDate
from django.http                     import FileResponse, HttpResponse
from .models                         import Product, Order, Category, OrderItem, ProductRating, AuditLog
from .serializers                    import (
    LoginSerializer, StatsSerializer,
    ProductSerializer, OrderSerializer, CategorySerializer,
    ProductRatingSerializer, AuditLogSerializer,
)
from .utils         import log_action
from .telegram_bot  import send_order_notification
import requests
import json
import io
from datetime import date, timedelta
from django.conf import settings
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

#fbv1
@api_view(['POST'])
@permission_classes([permissions.AllowAny])
def login_view(request):
    serializer = LoginSerializer(data=request.data)
    serializer.is_valid(raise_exception=True)
    
    user  = serializer.validated_data['user']
    token, _ = Token.objects.get_or_create(user=user)
    log_action(request, 'admin_login', 'User', user.id, f'Вход: {user.username}')
    return Response({
        'token':    token.key,
        'username': user.username,
    })

#fbv2
@api_view(['POST'])
def logout_view(request):
    request.user.auth_token.delete()
    return Response({'detail': 'Вы вышли из системы'})

#fbv3
@api_view(['GET'])
def dashboard_stats(request):
    from django.db.models import F
    
    total_revenue = Order.objects.filter(
        status__in=['confirmed', 'delivered']
    ).aggregate(total=Sum('total'))['total'] or 0

    total_orders = Order.objects.count()
    new_orders   = Order.objects.filter(status='new').count()

    low_stock_count = Product.objects.filter(stock__lte=5).count()

    top_products = (
        Product.objects
        .annotate(sold=Sum('order_items__quantity'))
        .filter(sold__isnull=False)
        .order_by('-sold')[:5]
        .values('id', 'name', 'sold')
    )

    data = {
        'total_revenue':   total_revenue,
        'total_orders':    total_orders,
        'new_orders':      new_orders,
        'low_stock_count': low_stock_count,
        'top_products':    list(top_products),
    }

    serializer = StatsSerializer(data)
    return Response(serializer.data)

# fbv 4 - ai search
@api_view(['POST'])
@permission_classes([permissions.AllowAny])
def ai_search(request):
    query = request.data.get('query', '').strip()
    
    if not query:
        return Response(
            {'error': 'Введите поисковый запрос'},
            status=status.HTTP_400_BAD_REQUEST
        )

    products = Product.objects.select_related('category').filter(stock__gt=0)
    products_list = [
        {
            'id':       p.id,
            'name':     p.name,
            'price':    float(p.price),
            'stock':    p.stock,
            'category': p.category.name if p.category else '',
            'description': p.description,
        }
        for p in products
    ]

    prompt = f"""
Ты — помощник в магазине. Тебе дан список товаров и запрос покупателя.
Верни ТОЛЬКО валидный JSON без лишнего текста, без markdown, без блоков кода.
Запрос покупателя: "{query}"
Список товаров:
{json.dumps(products_list, ensure_ascii=False)}
Ответ должен быть строго в таком формате:
{{
  "ids": [1, 2, 3],
  "message": "Короткий комментарий на русском почему эти товары подходят"
}}
Если ничего не подходит:
{{
  "ids": [],
  "message": "Не нашёл подходящих товаров"
}}
"""
    try:
        response = requests.post(
            f'https://generativelanguage.googleapis.com/v1beta/models/gemini-3-flash-preview:generateContent?key={settings.GEMINI_API_KEY}',
            headers={'Content-Type': 'application/json'},
            json={
                'contents': [{'parts': [{'text': prompt}]}],
                'generationConfig': {
                    'temperature': 0.3, 
                }
            },
            timeout=15
        )
        response.raise_for_status()
        data = response.json()
        raw_text = data['candidates'][0]['content']['parts'][0]['text'].strip()
        result = json.loads(raw_text)
        found_ids     = result.get('ids', [])
        found_products = Product.objects.filter(id__in=found_ids).select_related('category')
        return Response({
            'message':  result.get('message', ''),
            'products': ProductSerializer(found_products, many=True).data
        })

    except json.JSONDecodeError:
        return Response(
            {'error': 'Gemini вернул неожиданный формат'},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )
    except requests.exceptions.Timeout:
        return Response(
            {'error': 'Gemini не ответил вовремя'},
            status=status.HTTP_504_GATEWAY_TIMEOUT
        )
    except Exception as e:
        return Response(
            {'error': str(e)},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )
# fbv5 - ai image search
@api_view(['POST'])
@permission_classes([permissions.AllowAny])
def ai_search_by_image(request):
    image_base64 = request.data.get('image_base64', '').strip()

    if not image_base64:
        return Response(
            {'error': 'Поле image_base64 обязательно'},
            status=status.HTTP_400_BAD_REQUEST
        )

    products = Product.objects.select_related('category').filter(stock__gt=0)
    products_list = [
        {
            'id':          p.id,
            'name':        p.name,
            'price':       float(p.price),
            'category':    p.category.name if p.category else '',
            'description': p.description,
        }
        for p in products
    ]

    prompt = (
        f'Here is a list of products in JSON: {json.dumps(products_list, ensure_ascii=False)}. '
        'Look at the image and return ONLY valid JSON: '
        '{ "ids": [matching product ids], "message": "explanation in Russian" }. '
        'No markdown, no code blocks.'
    )

    try:
        response = requests.post(
            f'https://generativelanguage.googleapis.com/v1beta/models/gemini-3-flash-preview:generateContent?key={settings.GEMINI_API_KEY}',
            headers={'Content-Type': 'application/json'},
            json={
                'contents': [{
                    'parts': [
                        {'inline_data': {'mime_type': 'image/jpeg', 'data': image_base64}},
                        {'text': prompt},
                    ]
                }]
            },
            timeout=20
        )
        response.raise_for_status()
        data     = response.json()
        raw_text = data['candidates'][0]['content']['parts'][0]['text'].strip()
        result   = json.loads(raw_text)

        found_ids      = result.get('ids', [])
        found_products = Product.objects.filter(id__in=found_ids).select_related('category')
        return Response({
            'message':  result.get('message', ''),
            'products': ProductSerializer(found_products, many=True, context={'request': request}).data,
        })

    except json.JSONDecodeError:
        return Response(
            {'error': 'Gemini вернул неожиданный формат'},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )
    except requests.exceptions.Timeout:
        return Response(
            {'error': 'Gemini не ответил вовремя'},
            status=status.HTTP_504_GATEWAY_TIMEOUT
        )
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# fbv6 - track order by tracking code
@api_view(['GET'])
@permission_classes([permissions.AllowAny])
def track_order(request, tracking_code):
    try:
        order = Order.objects.prefetch_related('items__product').get(tracking_code=tracking_code)
    except Order.DoesNotExist:
        return Response({'error': 'Заказ не найден'}, status=status.HTTP_404_NOT_FOUND)
    return Response(OrderSerializer(order).data)


# fbv7 - sales chart last 30 days
@api_view(['GET'])
def sales_chart(request):
    today = date.today()
    start = today - timedelta(days=29)

    qs = (
        Order.objects
        .filter(status__in=['confirmed', 'delivered'], created_at__date__gte=start)
        .annotate(day=TruncDate('created_at'))
        .values('day')
        .annotate(revenue=Sum('total'))
        .order_by('day')
    )

    revenue_by_date = {row['day']: float(row['revenue']) for row in qs}

    result = []
    for i in range(30):
        day = start + timedelta(days=i)
        result.append({'date': day.isoformat(), 'revenue': revenue_by_date.get(day, 0.0)})

    return Response(result)


# fbv8 - export orders to excel
@api_view(['GET'])
def export_orders(request):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Заказы'

    headers = ['ID', 'Клиент', 'Телефон', 'Статус', 'Сумма (₸)', 'Дата', 'Состав заказа']
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid')

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill

    orders = Order.objects.prefetch_related('items__product').all()

    for row_idx, order in enumerate(orders, 2):
        composition = ', '.join(
            f'{item.product.name} x{item.quantity}'
            for item in order.items.all()
            if item.product
        )
        ws.cell(row=row_idx, column=1, value=order.id)
        ws.cell(row=row_idx, column=2, value=order.customer_name or '')
        ws.cell(row=row_idx, column=3, value=order.customer_phone or '')
        ws.cell(row=row_idx, column=4, value=order.get_status_display())
        ws.cell(row=row_idx, column=5, value=float(order.total))
        ws.cell(row=row_idx, column=6, value=order.created_at.strftime('%Y-%m-%d %H:%M'))
        ws.cell(row=row_idx, column=7, value=composition)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="orders.xlsx"'
    wb.save(response)
    return response


# fbv9 - rate product
@api_view(['POST'])
@permission_classes([permissions.AllowAny])
def rate_product(request, pk):
    try:
        product = Product.objects.get(pk=pk)
    except Product.DoesNotExist:
        return Response({'error': 'Товар не найден'}, status=status.HTTP_404_NOT_FOUND)

    try:
        score = int(request.data.get('score', 0))
    except (TypeError, ValueError):
        score = 0

    if not 1 <= score <= 5:
        return Response(
            {'error': 'Оценка должна быть от 1 до 5'},
            status=status.HTTP_400_BAD_REQUEST
        )

    if not request.session.session_key:
        request.session.create()

    ProductRating.objects.update_or_create(
        product=product,
        session_key=request.session.session_key,
        defaults={'score': score},
    )

    return Response(ProductSerializer(product, context={'request': request}).data)


# fbv10 - product recommendations
@api_view(['GET'])
@permission_classes([permissions.AllowAny])
def product_recommendations(request, pk):
    order_ids = OrderItem.objects.filter(product_id=pk).values_list('order_id', flat=True)

    co_products = (
        OrderItem.objects
        .filter(order_id__in=order_ids)
        .exclude(product_id=pk)
        .exclude(product__isnull=True)
        .values('product_id')
        .annotate(freq=Count('product_id'))
        .order_by('-freq')[:4]
    )
    found_ids = [row['product_id'] for row in co_products]

    if len(found_ids) < 4:
        cat_id = Product.objects.filter(pk=pk).values_list('category_id', flat=True).first()
        extra = list(
            Product.objects
            .filter(category_id=cat_id)
            .exclude(pk=pk)
            .exclude(pk__in=found_ids)
            .values_list('id', flat=True)[:4 - len(found_ids)]
        )
        found_ids += extra

    if len(found_ids) < 4:
        extra = list(
            Product.objects
            .exclude(pk=pk)
            .exclude(pk__in=found_ids)
            .order_by('?')
            .values_list('id', flat=True)[:4 - len(found_ids)]
        )
        found_ids += extra

    products = Product.objects.filter(pk__in=found_ids).select_related('category')
    return Response({
        'recommendations': ProductSerializer(products, many=True, context={'request': request}).data
    })


#cbv1
class ProductListCreateView(generics.ListCreateAPIView):
    serializer_class = ProductSerializer
    parser_classes   = [MultiPartParser, FormParser, JSONParser]

    def get_queryset(self):
        qs     = Product.objects.select_related('category').all()
        params = self.request.query_params

        if 'category' in params:
            qs = qs.filter(category_id=params['category'])
        if 'min_price' in params:
            qs = qs.filter(price__gte=params['min_price'])
        if 'max_price' in params:
            qs = qs.filter(price__lte=params['max_price'])
        if params.get('in_stock') == 'true':
            qs = qs.filter(stock__gt=0)

        ordering = params.get('ordering')
        if ordering in ('price', '-price', 'name', '-name', '-created_at'):
            qs = qs.order_by(ordering)

        return qs

    def get_permissions(self):
        if self.request.method == 'GET':
            return [permissions.AllowAny()]
        return [permissions.IsAuthenticated()]

    def get_serializer_context(self):
        return {'request': self.request}

    def create(self, request, *args, **kwargs):
        response = super().create(request, *args, **kwargs)
        log_action(request, 'product_created', 'Product', response.data.get('id'),
                   f"Создан товар: {response.data.get('name')}")
        return response

#cbv2
class ProductDetailView(generics.RetrieveUpdateDestroyAPIView):
    queryset         = Product.objects.all()
    serializer_class = ProductSerializer
    parser_classes   = [MultiPartParser, FormParser, JSONParser]

    def get_permissions(self):
        if self.request.method == 'GET':
            return [permissions.AllowAny()]
        return [permissions.IsAuthenticated()]

    def get_serializer_context(self):
        return {'request': self.request}

    def update(self, request, *args, **kwargs):
        response = super().update(request, *args, **kwargs)
        log_action(request, 'product_updated', 'Product', kwargs.get('pk'),
                   f"Обновлён товар #{kwargs.get('pk')}")
        return response

    def destroy(self, request, *args, **kwargs):
        pk = kwargs.get('pk')
        response = super().destroy(request, *args, **kwargs)
        log_action(request, 'product_deleted', 'Product', pk,
                   f"Удалён товар #{pk}")
        return response

#cbv3
class OrderListCreateView(generics.ListCreateAPIView):
    serializer_class = OrderSerializer

    def get_permissions(self):
        if self.request.method == 'POST':
            return [permissions.AllowAny()]
        return [permissions.IsAuthenticated()]

    def get_queryset(self):
        return Order.objects.prefetch_related('items__product').all()

    def perform_create(self, serializer):
        order = serializer.save()
        order.refresh_from_db()
        send_order_notification(order)


#cbv4
class OrderDetailView(generics.RetrieveUpdateAPIView):
    queryset         = Order.objects.prefetch_related('items__product').all()
    serializer_class = OrderSerializer

    def update(self, request, *args, **kwargs):
        response = super().update(request, *args, **kwargs)
        if 'status' in request.data:
            log_action(request, 'order_status_changed', 'Order', kwargs.get('pk'),
                       f"Статус заказа #{kwargs.get('pk')} изменён на: {request.data.get('status')}")
        return response


#cbv5
class CategoryListCreateView(generics.ListCreateAPIView):
    queryset         = Category.objects.all()
    serializer_class = CategorySerializer

    def get_permissions(self):
        if self.request.method == 'GET':
            return [permissions.AllowAny()]
        return [permissions.IsAuthenticated()]


#cbv6
class AuditLogPagination(PageNumberPagination):
    page_size = 50


class AuditLogListView(generics.ListAPIView):
    serializer_class = AuditLogSerializer
    pagination_class = AuditLogPagination

    def get_queryset(self):
        qs     = AuditLog.objects.select_related('user').all()
        entity = self.request.query_params.get('entity')
        if entity:
            qs = qs.filter(entity=entity)
        return qs
